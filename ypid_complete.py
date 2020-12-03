# /usr/bin/env python3
# -*- coding: utf8 -*-

import re
import xlrd
import openpyxl


ypid_dict = {}
code_re = '8.*'

# 读取采购平台文件的YPID
f_pt = xlrd.open_workbook(r'./采购平台中本院产品带HIS编码20191206发.xls')
jgcp_sheet = f_pt.sheet_by_name('机构产品管理报表')

for i in range(jgcp_sheet.nrows):
    row = jgcp_sheet.row_values(i)
    if re.match(code_re, row[0]):
        ypid_dict[row[0]] = row[jgcp_sheet.ncols - 1]
        # print(ypid_dict[row[0]])

# 读取毒麻药文件的YPID
f_dm = xlrd.open_workbook(r'./YPID麻醉药品、精神药品2019.12.10.xlsx')
for i in range(len(f_dm.sheets())):
    sheet = f_dm.sheet_by_index(i)
    for i in range(sheet.nrows):
        row = sheet.row_values(i)
        if re.match(code_re, str(row[0])):
            ypid_dict[row[0]] = row[sheet.ncols - 1]
            # print(ypid_dict[row[0]])

# print(ypid_dict['854359'])

# print(len(ypid_dict))

#############修改结果###############
# 先备份
# shutil.copyfile('./国卫办药政函【2019】834号要求2019-12-11.xlsx',
#                 './国卫办药政函【2019】834号要求2019-12-11-new.xlsx')
change_count = 0
f_re = openpyxl.load_workbook(filename='./国卫办药政函【2019】834号要求2019-12-11.xlsx')
for name in f_re.sheetnames:
    sheet = f_re[name]
    # import pdb
    # pdb.set_trace()
    if re.match('.*附表(1|5|6|7).*', sheet.title):
        for i in range(sheet.max_row):
            if re.match(code_re, str(sheet.cell(i+1, 2).value)):
                if str(sheet.cell(i+1, 2).value)[0:6] in ypid_dict.keys() and \
                    sheet.cell(i+1, 1).value != ypid_dict[str(sheet.cell(i+1, 2).value)[0:6]]:
                    sheet.cell(i+1, 1).value = ypid_dict[str(sheet.cell(i+1, 2).value)[0:6]]
                    change_count = change_count + 1
    # if re.match('.*附表7.*', sheet.title):
    #     for i in range(sheet.max_row):
    #         if re.match(code_re, str(sheet.cell(i+1, 3).value)):
    #             if str(sheet.cell(i+1, 3).value)[0:6] in ypid_dict.keys() and \
    #                 sheet.cell(i+1, 1).value != ypid_dict[str(sheet.cell(i+1, 3).value)[0:6]]:
    #                 sheet.cell(i+1, 1).value = ypid_dict[str(sheet.cell(i+1, 3).value)[0:6]]
    #                 change_count = change_count + 1
f_re.save('./国卫办药政函【2019】834号要求2019-12-11_new.xlsx')
print('%s changed' % change_count)

